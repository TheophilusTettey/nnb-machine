"""Read reference Excel workbooks (per-cell inclusion and compartment B)."""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl


def _cell_num(value) -> int | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value)
    return None


def read_manual_block(workbook_path: str | Path, block_header: str) -> list[dict]:
    """
    Parse 'N&B analysis' sheet block starting at row where column B == block_header.
    Returns per-cell manual records.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["N&B analysis"]
    start = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == block_header:
            start = row
            break
    if start is None:
        raise ValueError(f"Block '{block_header}' not found in {workbook_path}")

    records = []
    for row in range(start + 1, ws.max_row + 1):
        cell_id = _cell_num(ws.cell(row, 1).value)
        if cell_id is None:
            # stop at blank separator after data rows
            if ws.cell(row, 2).value in (None, "mean") and not records:
                continue
            if records:
                break
            continue

        b_nuc = ws.cell(row, 2).value
        e_nuc = ws.cell(row, 3).value
        norm_nuc = ws.cell(row, 4).value
        i_nuc_lo = ws.cell(row, 5).value
        i_nuc_hi = ws.cell(row, 6).value
        obs_nuc = ws.cell(row, 7).value
        b_arr = ws.cell(row, 9).value
        e_arr = ws.cell(row, 10).value
        norm_arr = ws.cell(row, 11).value
        i_arr_lo = ws.cell(row, 12).value
        i_arr_hi = ws.cell(row, 13).value
        obs_arr = ws.cell(row, 14).value

        records.append(
            {
                "cell_id": cell_id,
                "nucleus": {
                    "B": b_nuc,
                    "epsilon": e_nuc,
                    "norm_epsilon": norm_nuc,
                    "intensity_min": i_nuc_lo,
                    "intensity_max": i_nuc_hi,
                    "obs": obs_nuc,
                    "included": b_nuc is not None and e_nuc is not None,
                },
                "array": {
                    "B": b_arr,
                    "epsilon": e_arr,
                    "norm_epsilon": norm_arr,
                    "intensity_min": i_arr_lo,
                    "intensity_max": i_arr_hi,
                    "obs": obs_arr,
                    "included": b_arr is not None and e_arr is not None,
                },
            }
        )
    return records


def manual_file_index(filename: str) -> int | None:
    m = re.match(r"^(\d+)", Path(filename).name)
    return int(m.group(1)) if m else None